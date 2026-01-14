import os
import subprocess
import cv2
import calendar
from datetime import datetime, date, timedelta
from flask import Flask, render_template, request, redirect, url_for, session, flash, Response, send_file, jsonify
from models import db, Student, Attendance
from recognize_knn_attendance import recognize_frame
from sqlalchemy import text
from openpyxl import Workbook


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STUDENT_IMAGE_DIR = os.path.join(BASE_DIR, "student_images")
EMBEDDING_DIR = os.path.join(BASE_DIR, "embeddings")
os.makedirs(STUDENT_IMAGE_DIR, exist_ok=True)
os.makedirs(EMBEDDING_DIR, exist_ok=True)

app = Flask(__name__)
app.config["SECRET_KEY"] = "super_secret_key"
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///" + os.path.join(BASE_DIR, "attendance.db")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
db.init_app(app)

ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "admin123"

STREAMING = False
_initialized_today = False


# ---------------- LOGIN ----------------
def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if "logged_in" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


# ---------------- LANDING & LOGIN ----------------
@app.route("/")
def landing():
    return render_template("landing.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        if request.form["username"] == ADMIN_USERNAME and request.form["password"] == ADMIN_PASSWORD:
            session["logged_in"] = True
            return redirect(url_for("dashboard"))
        flash("Invalid Credentials")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("landing"))


# ---------------- DASHBOARD ----------------
@app.route("/dashboard")
@login_required
def dashboard():
    today = date.today()
    present_ids = {a.student_id for a in Attendance.query.filter_by(date=today, status="Present").all()}
    return render_template("dashboard.html", present=present_ids)


# ---------------- INITIALIZE TODAY ----------------
def initialize_today_attendance():
    global _initialized_today
    if _initialized_today:
        return

    today = date.today()
    students = Student.query.all()
    existing = {a.student_id for a in Attendance.query.filter_by(date=today).all()}

    for s in students:
        if s.id not in existing:
            record = Attendance(
                student_id=s.id,
                date=today,
                day_of_week=today.strftime("%A"),
                status="Absent",
                timestamp=datetime.now(),
                source="system"
            )
            db.session.add(record)

    db.session.commit()
    _initialized_today = True


# ---------------- MARK ATTENDANCE ----------------
def mark_attendance(student_name):
    today = date.today()

    student = Student.query.filter(db.func.lower(Student.full_name) == student_name.lower()).first()
    if not student:
        print("❌ Not found:", student_name)
        return

    rec = Attendance.query.filter_by(student_id=student.id, date=today).first()

    if rec:
        rec.status = "Present"
        rec.timestamp = datetime.now()
    else:
        db.session.add(Attendance(
            student_id=student.id,
            date=today,
            day_of_week=today.strftime("%A"),
            status="Present",
            timestamp=datetime.now(),
            source="webcam"
        ))

    db.session.commit()
    print("🟢 Marked Present:", student_name)


# ---------------- ADD STUDENT ----------------
@app.route("/students/add", methods=["GET", "POST"])
@login_required
def add_student():
    if request.method == "POST":

        full_name = request.form.get("full_name")
        roll_no = request.form.get("roll_no")
        images = request.files.getlist("images")

        # -----------------------------
        # NEW: Folder based on FULL NAME
        # -----------------------------
        folder = os.path.join(STUDENT_IMAGE_DIR, full_name)
        os.makedirs(folder, exist_ok=True)

        # Save images into the name folder
        for img in images:
            if img and img.filename:
                img.save(os.path.join(folder, img.filename))

        # Save student in DB
        s = Student(
            full_name=full_name,
            roll_no=roll_no,
            image_folder=folder
        )
        db.session.add(s)
        db.session.commit()

        # -----------------------------
        # Generate embeddings using NAME
        # -----------------------------
        try:
            subprocess.Popen([
                "python",
                os.path.join(BASE_DIR, "generate_embeddings_per_image.py"),
                "--input", folder,
                "--out", EMBEDDING_DIR,
                "--name", full_name   # IMPORTANT
            ])
        except Exception as e:
            print("Embedding error:", e)

        # -----------------------------
        # Train KNN on NAME labels
        # -----------------------------
        try:
            subprocess.Popen([
                "python",
                os.path.join(BASE_DIR, "train_knn.py")
            ])
        except Exception as e:
            print("KNN training error:", e)

        flash("Student added. Embeddings & KNN training started.")
        return redirect(url_for("students_list"))

    return render_template("add_student.html")

@app.route("/students/update")
@login_required
def select_student_update():
    students = Student.query.order_by(Student.full_name).all()
    return render_template("select_student_update.html", students=students)

@app.route("/students/update/<int:id>", methods=["GET", "POST"])
@login_required
def update_student(id):
    student = Student.query.get_or_404(id)

    if request.method == "POST":
        # Get updated values
        student.full_name = request.form.get("full_name")
        student.roll_no = request.form.get("roll_no")
        student.mobile = request.form.get("mobile")
        student.email = request.form.get("email")

        images = request.files.getlist("images")

        # Only update images if new ones are uploaded
        if images and images[0].filename != "":
            folder = student.image_folder
            os.makedirs(folder, exist_ok=True)

            for img in images:
                img.save(os.path.join(folder, img.filename))

            # OPTIONAL: retrigger embeddings + KNN
            subprocess.Popen([
                "python",
                os.path.join(BASE_DIR, "generate_embeddings_per_image.py")
            ])
            subprocess.Popen([
                "python",
                os.path.join(BASE_DIR, "train_knn.py")
            ])

        db.session.commit()
        flash("Student updated successfully")
        return redirect(url_for("students_list"))

    return render_template("update_student.html", student=student)


@app.route("/students")
@login_required
def students_list():
    return render_template("students_list.html", students=Student.query.all())

@app.route("/attendance/student_list")
@login_required
def attendance_student_list():
    students = Student.query.order_by(Student.roll_no).all()
    return render_template("attendance_student_list.html", students=students)

@app.route("/students/delete/<int:id>")
@login_required
def delete_student(id):
    s = Student.query.get_or_404(id)
    db.session.delete(s)
    db.session.commit()
    flash("Student Deleted")
    return redirect(url_for("students_list"))


# ---------------- WEBCAM PAGE ----------------
@app.route("/start_webcam")
@login_required
def start_webcam():
    return render_template("start_webcam.html")


@app.route("/webcam")
@login_required
def webcam():
    return render_template("start_webcam.html")


@app.route("/start_stream")
@login_required
def start_stream():
    global STREAMING
    STREAMING = True
    return jsonify({"status": "started"})


@app.route("/stop_stream")
@login_required
def stop_stream():
    global STREAMING
    STREAMING = False
    return jsonify({"status": "stopped"})


def gen_frames():
    global STREAMING
    cap = cv2.VideoCapture(0)

    with app.app_context():
        initialize_today_attendance()

    while STREAMING:
        success, frame = cap.read()
        if not success:
            break

        processed, name = recognize_frame(frame)

        if name:
            with app.app_context():
                mark_attendance(name)

        ret, buffer = cv2.imencode('.jpg', processed)
        yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')

    cap.release()


@app.route("/video_feed")
@login_required
def video_feed():
    return Response(gen_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')


# ---------------- TODAY'S ATTENDANCE ----------------
@app.route("/attendance/today")
@login_required
def attendance_today():
    today = date.today()
    students = Student.query.order_by(Student.roll_no).all()
    records = {a.student_id: a.status for a in Attendance.query.filter_by(date=today).all()}
    return render_template("attendance_today.html", students=students, records=records)


@app.route("/submit_today", methods=["POST"])
@login_required
def submit_today():
    initialize_today_attendance()
    flash("Today's attendance saved.")
    return redirect(url_for("attendance_today"))


@app.route("/export_today")
@login_required
def export_today():
    today = date.today()

    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Roll No", "Date", "Day", "Status"])

    students = Student.query.order_by(Student.roll_no).all()
    records = {r.student_id: r.status for r in Attendance.query.filter_by(date=today).all()}

    for s in students:
        ws.append([s.full_name, s.roll_no, str(today), today.strftime("%A"), records.get(s.id, "Absent")])

    out_file = f"today_attendance.xlsx"
    wb.save(out_file)

    return send_file(out_file, as_attachment=True)


# ---------------- MONTHLY ATTENDANCE ----------------
@app.route("/attendance/monthly")
@login_required
def attendance_monthly():
    today = date.today()
    year, month = today.year, today.month
    _, days = calendar.monthrange(year, month)

    summary = {}
    for d in range(1, days + 1):
        day_date = date(year, month, d)
        present = Attendance.query.filter_by(date=day_date, status="Present").count()
        absent = Attendance.query.filter_by(date=day_date, status="Absent").count()
        summary[d] = {"present": present, "absent": absent}

    return render_template("attendance_monthly.html",
                           days=range(1, days + 1),
                           summary=summary,
                           year=year,
                           month=month,
                           date=date)


# ---------------- DAY ATTENDANCE ----------------
@app.route("/attendance/day/<int:day>")
@login_required
def attendance_day(day):
    today = date.today()
    selected = date(today.year, today.month, day)

    sql = text("""
        SELECT s.full_name, s.roll_no, a.status
        FROM attendance a
        JOIN student s ON a.student_id = s.id
        WHERE a.date = :d
        ORDER BY s.roll_no
    """)

    rows = db.session.execute(sql, {"d": selected}).fetchall()

    records = [(r[0], r[1], r[2]) for r in rows]

    return render_template("attendance_day.html", date=selected, records=records)



@app.route("/export_day/<int:day>")
@login_required
def export_day(day):
    today = date.today()
    selected = date(today.year, today.month, day)

    sql = text("""
        SELECT s.full_name, s.roll_no, a.status
        FROM attendance a
        JOIN student s ON a.student_id = s.id
        WHERE a.date = :d
        ORDER BY s.roll_no
    """)

    rows = db.session.execute(sql, {"d": selected}).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendance {selected}"

    ws.append(["Name", "Roll No", "Status"])

    for name, roll, status in rows:
        ws.append([name, roll, status])

    filename = f"attendance_{selected}.xlsx"
    wb.save(filename)

    return send_file(filename, as_attachment=True)

@app.route("/attendance/student/<int:student_id>")
@login_required
def attendance_student_month(student_id):
    student = Student.query.get_or_404(student_id)

    today = date.today()
    year = today.year
    month = today.month

    # Fetch all attendance records for this student for this month
    records = Attendance.query.filter(
        Attendance.student_id == student_id,
        Attendance.date.between(date(year, month, 1), date(year, month, 31))
    ).order_by(Attendance.date).all()

    # --- SUMMARY CALCULATIONS ---
    total_days = len(records)
    present_days = sum(1 for r in records if r.status == "Present")
    absent_days = sum(1 for r in records if r.status == "Absent")

    percentage = 0
    if total_days > 0:
        percentage = round((present_days / total_days) * 100, 2)

    summary = {
        "total_days": total_days,
        "present": present_days,
        "absent": absent_days,
        "percentage": percentage
    }

    return render_template(
        "attendance_student_month.html",
        student=student,
        records=records,
        year=year,
        month=month,
        summary=summary
    )

@app.route("/export_student_month/<int:student_id>")
@login_required
def export_student_month(student_id):
    student = Student.query.get_or_404(student_id)

    today = date.today()
    year = today.year
    month = today.month

    # Get attendance records
    records = Attendance.query.filter(
        Attendance.student_id == student_id,
        Attendance.date.between(date(year, month, 1), date(year, month, 31))
    ).order_by(Attendance.date).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f"{student.full_name} Attendance"

    ws.append(["Date", "Day", "Status"])

    for r in records:
        ws.append([str(r.date), r.day_of_week, r.status])

    filename = f"{student.full_name}_attendance_{month}_{year}.xlsx"
    filepath = os.path.join(BASE_DIR, filename)
    wb.save(filepath)

    return send_file(filepath, as_attachment=True)


def calculate_attendance_percent(student_id):
    today = date.today()
    year, month = today.year, today.month
    start_date = date(year, month, 1)
    end_date = date(year, month, calendar.monthrange(year, month)[1])

    # Count present days
    present_count = Attendance.query.filter(
        Attendance.student_id == student_id,
        Attendance.date.between(start_date, end_date),
        Attendance.status == "Present"
    ).count()

    # Count total weekdays in the month
    total_weekdays = 0
    current = start_date
    while current <= end_date:
        if current.weekday() < 5:  # Monday to Friday
            total_weekdays += 1
        current += timedelta(days=1)

    if total_weekdays == 0:
        return 0
    return round((present_count / total_weekdays) * 100, 2)


def generate_excel_for_student(student_id):
    student = Student.query.get_or_404(student_id)
    today = date.today()
    year, month = today.year, today.month

    # Get attendance records
    records = Attendance.query.filter(
        Attendance.student_id == student_id,
        Attendance.date.between(date(year, month, 1), date(year, month, 31))
    ).order_by(Attendance.date).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f"{student.full_name} Attendance"

    ws.append(["Date", "Day", "Status"])

    for r in records:
        ws.append([str(r.date), r.day_of_week, r.status])

    filename = f"{student.full_name}_attendance_{month}_{year}.xlsx"
    filepath = os.path.join(BASE_DIR, filename)
    wb.save(filepath)

    return filepath


@app.route("/notify_monthly", methods=["POST"])
@login_required
def notify_monthly():
    students = Student.query.all()

    for s in students:
        percent = calculate_attendance_percent(s.id)
        excel_path = generate_excel_for_student(s.id)

        notify_student(s, percent, excel_path)

    flash("Monthly notifications sent successfully")
    return redirect(url_for("attendance_monthly"))

@app.route("/get_attendance_count")
@login_required
def get_attendance_count():
    try:
        today = date.today()
        # This counts unique students who are marked "Present" today
        count = Attendance.query.filter_by(date=today, status="Present").count()
        return jsonify({"success": True, "count": count})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
    
# ---------------- INIT ----------------
if __name__ == "__main__":
    with app.app_context():
        db.create_all()
    app.run(debug=True)
